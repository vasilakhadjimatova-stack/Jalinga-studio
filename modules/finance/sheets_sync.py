"""Google Sheets («Jalinga 2026») ↔ moliya bazasi sinxronizatsiyasi.

Ikki yo'l:
1. **Jonli sync** — jadval havola orqali ochiq, shuning uchun autentifikatsiyasiz
   xlsx eksport URL'idan yuklab olamiz (kredensial talab qilinmaydi):
   https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx
2. **Snapshot** — data/finance_snapshot.json (repoga qo'shilgan nusxa) — birinchi
   ishga tushishda internet/ulanishsiz ham ma'lumot bo'lishi uchun.

Sync qoidasi: source='sheet' yozuvlar to'liq qayta yoziladi, qo'lda
kiritilganlar (source='manual') saqlanadi.
"""
import io
import json
import logging
import os
from datetime import datetime, date

from database import db
from models.finance import (FinWallet, FinCategory, FinTransaction,
                            FinDebt, FinSetting)

logger = logging.getLogger(__name__)

SNAPSHOT_PATH = os.path.join(os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))), "data", "finance_snapshot.json")

# Sheets'dagi ruscha qiymatlar → ichki kodlar
DIRECTION_MAP = {"Поступление": "in", "Выбытие": "out"}
ACTIVITY_MAP = {
    "Операционная": "operating",
    "Инвестиционная": "investing",
    "Финансовая": "financing",
    "Техническая операция": "technical",
}
ACTIVITY_LABELS = {
    "operating": "Operatsion faoliyat",
    "investing": "Investitsion faoliyat",
    "financing": "Moliyaviy faoliyat",
    "technical": "Texnik (hisoblar aro o'tkazma)",
}

# Касса varag'idagi blok nomi → jurnal (ДДС данные) hamyon nomi.
# Prefiks bo'yicha solishtiriladi (Sheets'da nom biroz farq qilishi mumkin).
KASSA_WALLET_MAP = [
    ("РС Jalinga", "РС Jalinga", "UZS"),
    ("карта 1466", "карта 1466 (3781)", "UZS"),
    ("карта 9933", "карта 9933", "UZS"),
    ("касса (сом)", "Наличные", "UZS"),
    ("касса (долл)", "$", "USD"),
]


def _iso(value):
    """Sana → 'YYYY-MM-DD' (datetime/date/str hammasi qabul qilinadi)."""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return str(value or "").strip()[:24]


def parse_workbook(xlsx_bytes, usd_rate=12000.0):
    """xlsx (bytes) → {'wallets','categories','transactions','debts'} dict.

    Sheets tuzilishi o'zgarsa ham iloji boricha yiqilmasin: qator darajasida
    xatolar tashlab ketiladi (log bilan).
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)

    out = {"wallets": [], "categories": [], "transactions": [], "debts": []}

    # ── Справочники → statyalar ─────────────────────────────────────────────
    if "Справочники" in wb.sheetnames:
        for i, row in enumerate(wb["Справочники"].iter_rows(min_row=3, max_col=3,
                                                            values_only=True)):
            name, group, act = (row + (None,) * 3)[:3]
            name = str(name or "").strip()
            direction = DIRECTION_MAP.get(str(group or "").strip())
            activity = ACTIVITY_MAP.get(str(act or "").strip())
            if name and direction and activity:
                out["categories"].append(dict(name=name, direction=direction,
                                              activity=activity, sort=i))

    # ── ДДС данные → tranzaksiyalar ────────────────────────────────────────
    if "ДДС данные" in wb.sheetnames:
        for row in wb["ДДС данные"].iter_rows(min_row=3, max_col=11,
                                              values_only=True):
            m, y, d, amount, wallet, ctr, purpose, sub, cat, grp, act = \
                (row + (None,) * 11)[:11]
            try:
                amount = float(amount)
            except (TypeError, ValueError):
                continue
            cat = str(cat or "").strip()
            direction = DIRECTION_MAP.get(str(grp or "").strip())
            activity = ACTIVITY_MAP.get(str(act or "").strip())
            if not cat or not direction or not activity or amount == 0:
                continue
            iso = _iso(d)
            try:
                year, month = int(y), int(m)
            except (TypeError, ValueError):
                if len(iso) >= 7 and iso[:4].isdigit():
                    year, month = int(iso[:4]), int(iso[5:7])
                else:
                    continue
            out["transactions"].append(dict(
                date=iso, year=year, month=month, amount=amount,
                wallet=str(wallet or "").strip() or "—",
                counterparty=str(ctr or "").strip(),
                purpose=str(purpose or "").strip(),
                subcategory=str(sub or "").strip(),
                category=cat, direction=direction, activity=activity))

    # ── Касса → ochilish qoldiqlari (yilning 1-kuni, B ustuni) ─────────────
    year_start = None
    if "Касса" in wb.sheetnames:
        ws = wb["Касса"]
        rows = list(ws.iter_rows(min_row=1, max_col=2, values_only=True))
        first_date = rows[0][1] if rows and len(rows[0]) > 1 else None
        if isinstance(first_date, (datetime, date)):
            year_start = first_date.year
        block = ""
        for a, b in rows:
            label = str(a or "").strip()
            if not label:
                continue
            if label != "остатки на начало дня":
                block = label
                continue
            try:
                opening = float(b or 0)
            except (TypeError, ValueError):
                opening = 0.0
            for prefix, wallet_name, cur in KASSA_WALLET_MAP:
                if block.lower().startswith(prefix.lower()):
                    if cur == "USD":     # $ kassa so'm ekvivalentida yuritiladi
                        opening *= usd_rate
                    out["wallets"].append(dict(
                        name=wallet_name, opening_balance=round(opening, 2),
                        opening_year=year_start or 2026, currency=cur,
                        sort=len(out["wallets"])))
                    break

    # Jurnalda uchragan, Kassada yo'q hamyonlar — 0 ochilish bilan qo'shiladi
    known = {w["name"] for w in out["wallets"]}
    for t in out["transactions"]:
        if t["wallet"] not in known:
            known.add(t["wallet"])
            out["wallets"].append(dict(name=t["wallet"], opening_balance=0.0,
                                       opening_year=year_start or 2026,
                                       currency="UZS", sort=len(out["wallets"])))
    # Jurnalda bor, Справочникда yo'q statyalar
    known_cats = {c["name"] for c in out["categories"]}
    for t in out["transactions"]:
        if t["category"] not in known_cats:
            known_cats.add(t["category"])
            out["categories"].append(dict(name=t["category"],
                                          direction=t["direction"],
                                          activity=t["activity"],
                                          sort=len(out["categories"])))

    # ── DOLG → qarzlar ─────────────────────────────────────────────────────
    if "DOLG" in wb.sheetnames:
        for row in wb["DOLG"].iter_rows(min_row=4, max_col=10, values_only=True):
            _, num, sana, debtor, creditor, reason, summa, repaid, _qold, rdate = \
                (row + (None,) * 10)[:10]
            try:
                amount = float(summa)
            except (TypeError, ValueError):
                continue
            # Jami-qatorlar (debitor/kreditor bo'sh) qarz emas — tashlab ketiladi
            if not str(debtor or "").strip() and not str(creditor or "").strip():
                continue
            try:
                repaid_f = float(repaid or 0)
            except (TypeError, ValueError):
                repaid_f = 0.0
            out["debts"].append(dict(
                date=_iso(sana), debtor=str(debtor or "").strip(),
                creditor=str(creditor or "").strip(),
                reason=str(reason or "").strip(), amount=amount,
                repaid=repaid_f, repaid_date=_iso(rdate)))

    wb.close()
    return out


def apply_parsed(data, source="sheet"):
    """Parse natijasini bazaga yozish. Sheet yozuvlari qayta yoziladi,
    manual'lar tegilmaydi. Commit chaqiruvchi zimmasida EMAS — shu yerda."""
    # Hamyonlar: upsert (ochilish qoldig'i yangilanadi)
    for w in data.get("wallets", []):
        row = FinWallet.query.filter_by(name=w["name"]).first()
        if row is None:
            row = FinWallet(name=w["name"])
            db.session.add(row)
        row.opening_balance = w.get("opening_balance", 0.0)
        row.opening_year = w.get("opening_year", 2026)
        row.currency = w.get("currency", "UZS")
        row.sort = w.get("sort", 0)
    # Statyalar: upsert
    for c in data.get("categories", []):
        row = FinCategory.query.filter_by(name=c["name"]).first()
        if row is None:
            row = FinCategory(name=c["name"])
            db.session.add(row)
        row.direction = c["direction"]
        row.activity = c["activity"]
        row.sort = c.get("sort", 0)
    # Tranzaksiyalar va qarzlar: sheet'dagilar to'liq almashtiriladi
    FinTransaction.query.filter_by(source=source).delete()
    for t in data.get("transactions", []):
        db.session.add(FinTransaction(source=source, **t))
    FinDebt.query.filter_by(source=source).delete()
    for d in data.get("debts", []):
        db.session.add(FinDebt(source=source, **d))
    FinSetting.set("last_sync", datetime.now().strftime("%Y-%m-%d %H:%M"))
    db.session.commit()
    return {"transactions": len(data.get("transactions", [])),
            "debts": len(data.get("debts", [])),
            "wallets": len(data.get("wallets", []))}


# Snapshot bo'lmaganда (dastur-native yangi o'rnatma) boshlang'ich seed —
# hisoblar va ДДС statyalari. Shusiz foydalanuvchi tranzaksiya qo'sha olmaydi
# (statya kerak). Admin keyin Sozlamalarда o'zgartiradi.
# Nomlar studiya-ulash (studio_link.METHOD_WALLET) bilan mos — mijoz to'lovi
# to'g'ri hisobga tushishi uchun.
DEFAULT_WALLETS = [
    ("РС Jalinga", "UZS"), ("карта 9933", "UZS"), ("Наличные", "UZS"),
    ("$", "USD"),
]
DEFAULT_CATEGORIES = [
    ("Поступление от клиента (запись)", "in", "operating"),
    ("Поступление от клиента (Вебинар)", "in", "operating"),
    ("Поступление от монтажа", "in", "operating"),
    ("Прочие поступления", "in", "operating"),
    ("зарплата", "out", "operating"),
    ("премия", "out", "operating"),
    ("аренда", "out", "operating"),
    ("аренда студии", "out", "operating"),
    ("налог АОС/дивиденд/Зп", "out", "operating"),
    ("Комиссия Банк/ комиссия плат. систем", "out", "operating"),
    ("Комунальные услуги", "out", "operating"),
    ("Ремонт", "out", "operating"),
    ("Абонентские подписки", "out", "operating"),
    ("организационные расходы", "out", "operating"),
    ("прочие расходы", "out", "operating"),
    ("Продажа ОС", "in", "investing"),
    ("Покупка ОС", "out", "investing"),
    ("Займ от собственника", "in", "financing"),
    ("Погашение тела кредита, займа", "out", "financing"),
    ("Дивиденды", "out", "financing"),
    ("Доход — Перевод между счетами", "in", "technical"),
    ("Расход — Перевод между счетами", "out", "technical"),
]


def seed_default_finance():
    """Bo'sh bazada (snapshot yo'q) boshlang'ich hisoblar + statyalar."""
    from models.finance import FinWallet, FinCategory
    if FinCategory.query.first() is None:
        for i, (name, d, act) in enumerate(DEFAULT_CATEGORIES):
            db.session.add(FinCategory(name=name, direction=d, activity=act,
                                       sort=i))
    if FinWallet.query.first() is None:
        for i, (name, cur) in enumerate(DEFAULT_WALLETS):
            db.session.add(FinWallet(name=name, currency=cur, sort=i,
                                     opening_year=datetime.now().year))
    db.session.commit()
    logger.info("Moliya: boshlang'ich hisoblar va statyalar seed qilindi")


def seed_default_recurring():
    """Kalendar bo'sh bo'lmasin: Jalinga'ning ma'lum oylik to'lovlari
    (jadval tarixidan olingan tipik qiymatlar). Faqat hech qanday doimiy
    to'lov bo'lmaganда bir marta seed qilinadi — admin keyin tahrirlaydi."""
    from models.finance import FinRecurring
    if FinRecurring.query.first() is not None:
        return
    defaults = [
        ("Ofis ijarasi", 6100000, 2, "аренда"),
        ("Abonent obunalar", 800000, 1, "Абонентские подписки"),
    ]
    for i, (name, amount, day, cat) in enumerate(defaults):
        db.session.add(FinRecurring(name=name, amount=amount, pay_day=day,
                                    category=cat, wallet="РС Jalinga", sort=i))
    db.session.commit()
    logger.info("Kalendar: %d ta doimiy to'lov seed qilindi", len(defaults))


def import_snapshot_if_empty():
    """Birinchi ishga tushish uchun boshlang'ich ma'lumot.

    • Baza to'la bo'lsa — hech narsa qilmaydi (faqat recurring seed).
    • Snapshot fayli bo'lsa (lokal dev / qo'lда tiklash) — undan yuklaydi.
    • Aks holda — default hisoblar + statyalarni seed qiladi (dastur-native).

    DIQQAT: snapshot fayli repoда SAQLANMAYDI (real moliya ma'lumot — maxfiy).
    Prod ma'lumoti Postgres'да yashaydi; bu funksiya faqat bo'sh bazaда ishlaydi.
    """
    if FinTransaction.query.first() is not None:
        seed_default_recurring()
        return None
    if not os.path.exists(SNAPSHOT_PATH):
        seed_default_finance()
        seed_default_recurring()
        return None
    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        data = json.load(f)
    stats = apply_parsed(data)
    FinSetting.set("last_sync", data.get("meta", {}).get("generated_at", ""))
    FinSetting.set("sync_source", "snapshot")
    db.session.commit()
    seed_default_recurring()
    logger.info(f"Moliya snapshot yuklandi: {stats}")
    return stats
